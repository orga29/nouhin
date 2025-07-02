#!/usr/bin/env python3
"""nouhin.py (2025‑07‑02 Streamlit 対応版)
在庫集計表 (.xlsm) から "納品数シート" を抽出・整形して保存するコアモジュール。
旧 v07z と同一仕様。

エクスポート関数 `prepare(src_path, delivery_date)` を追加し、
Streamlit などから呼び出しやすくした。
"""
from __future__ import annotations

import sys
import shutil
from datetime import datetime
from pathlib import Path
from copy import copy as style_copy
from string import ascii_uppercase

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import PatternFill

# ---------------- 定数 ----------------
SRC_SHEET = "在庫集計表"
DST_SHEET = "納品数"
DATA_START_ROW_SRC = 8
DATA_START_ROW_DST = 5
FIXED_COLS = ["A", "B", "AN", "AQ"]
COL_RENAME_MAP = {"A": "商品コード", "B": "商品名", "AN": "箱/こ/不", "AQ": "荷受"}
WEEKDAY_RANGE = {0:("H","M"),1:("M","R"),2:("R","W"),3:("W","AB"),4:("AB","AG"),5:("AG","AL")}

# ---------------- Utilities ----------------

def col_range(a: str, b: str):
    return [get_column_letter(i) for i in range(column_index_from_string(a), column_index_from_string(b)+1)]


def load_df(path: Path, letters: list[str]):
    df = pd.read_excel(path, sheet_name=SRC_SHEET, header=None, names=letters,
                       usecols=",".join(letters), skiprows=DATA_START_ROW_SRC-1)
    df.rename(columns=COL_RENAME_MAP, inplace=True)
    df.rename(columns={letters[-1]: "納品数"}, inplace=True)
    return df[df["商品コード"].notna()].reset_index(drop=True)


def sort_final(df: pd.DataFrame):
    I = df.iloc[:, 8].fillna('').astype(str).str.strip()
    rank = np.where(I == '箱', 3, 2)
    pref = df['商品名'].astype(str).str[0]
    sub1 = np.select([(rank == 2) & (pref == '■'), (rank == 2) & (pref == '□')], [0, 1], default=2)
    sub1 = np.where(rank == 3, 3, sub1)
    sub2 = np.where(sub1 >= 1, df['商品コード'].astype(str), '')
    return (df.assign(_r=rank, _s1=sub1, _s2=sub2)
              .sort_values(['_r', '_s1', '_s2'], ascending=[False, True, True], kind='mergesort')
              .drop(columns=['_r','_s1','_s2']).reset_index(drop=True))

# --- Excel helper ---

def clear_from_row(ws: Worksheet, row: int):
    for r in range(row, ws.max_row + 1):
        ws.row_dimensions[r].height = None
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = None
            cell.border = None
            cell.fill = PatternFill(fill_type=None)
            cell.number_format = "General"
            cell.alignment = None


def copy_style_row(ws: Worksheet, src: int, dst: int):
    for c in range(1, ws.max_column + 1):
        sc, dc = ws.cell(row=src, column=c), ws.cell(row=dst, column=c)
        dc.font = style_copy(sc.font)
        dc.border = style_copy(sc.border)
        dc.fill = style_copy(sc.fill)
        dc.number_format = sc.number_format
        dc.alignment = style_copy(sc.alignment)


def paste_df(ws: Worksheet, df: pd.DataFrame, start: int):
    base_h = ws.row_dimensions[start].height
    need_rows = len(df)
    if need_rows > (ws.max_row - start + 1):
        ws.insert_rows(ws.max_row + 1, need_rows - (ws.max_row - start + 1))
    for i in range(need_rows):
        r = start + i
        ws.row_dimensions[r].height = base_h
        copy_style_row(ws, start, r)
    for ridx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start):
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)
    clear_from_row(ws, start + need_rows)

# ------------ File helper ------------

def safe_copy(src: Path) -> Path:
    base = src.with_name(src.stem + "_prepared.xlsm")
    if not base.exists():
        shutil.copyfile(src, base)
        return base
    for i in range(1, 100):
        cand = src.with_name(f"{src.stem}_prepared({i}).xlsm")
        try:
            shutil.copyfile(src, cand)
            return cand
        except PermissionError:
            continue
    raise PermissionError("出力ファイルを作成できません — 既存ファイルを閉じてください")

# ------------- Write -------------

def write_out(src: Path, df: pd.DataFrame) -> Path:
    out = safe_copy(src)
    wb  = load_workbook(out, keep_vba=True)
    ws  = wb[DST_SHEET]

    paste_df(ws, df, DATA_START_ROW_DST)

    for i in range(len(df)):
        r = DATA_START_ROW_DST + i
        ws[f'K{r}'].value = f'=C{r}-E{r}+G{r}'
        ws[f'M{r}'].value = (
            f'=IF(COUNTIFS($B{r},"*東一*")>0,"東一",'
            f'IF(COUNTIF($L{r},"■")>0,'
            f'IF(COUNTIF($J{r},"○")>0,"荷受",$K{r}*-1),""))')

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth   = 1
    ws.page_margins            = PageMargins(top=0.4, bottom=0.4, left=0.3, right=0.3)
    ws.print_options.horizontalCentered = True

    last_row = DATA_START_ROW_DST + len(df) - 1
    ws.print_area = [f'A1:M{last_row}']

    wb.save(out)
    return out

# -------- prepare(): Streamlit 用入口 --------

def prepare(src_path: str | Path, delivery_date: str) -> str:
    """Streamlit / pytest から直接呼び出すラッパー。"""
    dt = datetime.strptime(delivery_date, "%Y-%m-%d")
    weekday = dt.weekday()
    if weekday == 6:
        raise ValueError("日曜は出荷対象外です")
    if weekday not in WEEKDAY_RANGE:
        raise ValueError("曜日マッピング未設定です")

    src = Path(src_path)
    if not src.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {src}")

    letters = FIXED_COLS + col_range(*WEEKDAY_RANGE[weekday])
    df = load_df(src, letters)
    df = df[~((df.iloc[:, 4] == 0) & (df.iloc[:, 6] == 0))]

    return str(write_out(src, sort_final(df)))

# ------------- CLI main -------------

def main(argv=None):
    args = sys.argv if argv is None else argv
    if len(args) < 3:
        print("Usage: python nouhin.py YYYY-MM-DD 集計表.xlsm", file=sys.stderr)
        sys.exit(1)

    delivery_date, src_path = args[1], args[2]
    out_path = prepare(src_path, delivery_date)
    print("✅ 出力ファイル:", out_path)


if __name__ == "__main__":
    main()
